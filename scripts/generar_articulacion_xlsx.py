#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Exporta la matriz de articulación a Excel para revisión/validación humana (pedido de Tellys).
Hojas: Políticas de Estado (arriba), Programas Presupuestales (abajo), Plan de Gobierno (Keiko),
y Resumen. Cada fila trae Tipo propuesto + Justificación + columnas para VALIDAR
(¿Correcto? / Tipo corregido / Observación) con menús desplegables.

Uso:  python3 scripts/generar_articulacion_xlsx.py
Salida: entregables/articulacion.xlsx
"""
import os, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = os.path.join(os.path.dirname(__file__), "..")
DATA = os.path.join(ROOT, "data")
OUT = os.path.join(ROOT, "entregables", "articulacion.xlsx")


def L(fn):
    return json.load(open(os.path.join(DATA, fn), encoding="utf-8"))


HEAD = Font(bold=True, color="FFFFFF", size=11)
HEADFILL = PatternFill("solid", fgColor="1F2A44")
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
THIN = Border(bottom=Side(style="thin", color="DDDDDD"))
TIPOS = '"igual_similar,similar,desagregado,causal"'
SINO = '"Sí,No,Revisar"'


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEAD; cell.fill = HEADFILL; cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def add_validation(ws, col_ok, col_tipo, nrows):
    dv_ok = DataValidation(type="list", formula1=SINO, allow_blank=True)
    dv_tipo = DataValidation(type="list", formula1=TIPOS, allow_blank=True)
    ws.add_data_validation(dv_ok); ws.add_data_validation(dv_tipo)
    dv_ok.add(f"{col_ok}2:{col_ok}{nrows+1}")
    dv_tipo.add(f"{col_tipo}2:{col_tipo}{nrows+1}")


def main():
    coms = {c["comision_id"]: c for c in L("articulacion.json")["articulaciones"]}
    an = L("acuerdo_nacional.json")
    pol_eje = {p["n"]: e["nombre"] for e in an["ejes"] for p in e["politicas"]}
    wb = Workbook()

    # ---- Hoja 1: Comisiones ↔ Políticas de Estado (Acuerdo Nacional)
    ws = wb.active; ws.title = "Políticas de Estado"
    cols = ["Comisión", "Eje del Acuerdo Nacional", "N° Política", "Política de Estado",
            "Tipo propuesto (IA)", "Justificación", "¿Correcto?", "Tipo corregido", "Observación"]
    ws.append(cols)
    n = 0
    for cid, a in sorted(coms.items(), key=lambda x: x[1]["comision_nombre"]):
        for x in a.get("acuerdo_nacional", []):
            ws.append([a["comision_nombre"], pol_eje.get(x["politica"], ""), x["politica"],
                       x.get("politica_nombre", ""), x["tipo"], x.get("justificacion", ""), "", "", ""])
            n += 1
    style_header(ws, len(cols)); add_validation(ws, "G", "H", n)
    for col, w in zip("ABCDEFGHI", [26, 34, 9, 40, 16, 70, 11, 16, 30]):
        ws.column_dimensions[col].width = w
    for r in range(2, n + 2):
        ws.cell(row=r, column=6).alignment = WRAP; ws.cell(row=r, column=4).alignment = WRAP

    # ---- Hoja 1b: Comisiones ↔ Objetivos del Acuerdo Nacional (nivel-3)
    nobj = 0
    try:
        objdata = L("articulacion_objetivos.json")["articulacion"]
        comnom = {cid: a["comision_nombre"] for cid, a in coms.items()}
        wso = wb.create_sheet("Objetivos AN")
        colso = ["Comisión", "N° Política", "Política de Estado", "Objetivo (letra)", "Objetivo",
                 "Tipo propuesto (IA)", "Justificación", "¿Correcto?", "Tipo corregido", "Observación"]
        wso.append(colso)
        for cid in sorted(objdata, key=lambda k: comnom.get(k, k)):
            for o in objdata[cid]:
                wso.append([comnom.get(cid, cid), o["politica"], o.get("politica_nombre", ""), o.get("letra", ""),
                            o.get("objetivo_texto", ""), o["tipo"], o.get("justificacion", ""), "", "", ""])
                nobj += 1
        style_header(wso, len(colso)); add_validation(wso, "H", "I", nobj)
        for col, w in zip("ABCDEFGHIJ", [24, 9, 32, 12, 52, 15, 55, 11, 15, 26]):
            wso.column_dimensions[col].width = w
        for r in range(2, nobj + 2):
            for cc in (3, 5, 7): wso.cell(row=r, column=cc).alignment = WRAP
    except Exception:
        pass

    # ---- Hoja 2: Comisiones ↔ Programas Presupuestales (MEF)
    ws2 = wb.create_sheet("Programas Presupuestales")
    cols2 = ["Comisión", "Código PP", "Programa Presupuestal", "Tipo propuesto (IA)",
             "Justificación", "¿Correcto?", "Tipo corregido", "Observación"]
    ws2.append(cols2); n2 = 0
    for cid, a in sorted(coms.items(), key=lambda x: x[1]["comision_nombre"]):
        for x in a.get("programas_presupuestales", []):
            ws2.append([a["comision_nombre"], x.get("codigo", ""), x.get("pp_nombre", ""),
                        x["tipo"], x.get("justificacion", ""), "", "", ""])
            n2 += 1
    style_header(ws2, len(cols2)); add_validation(ws2, "F", "G", n2)
    for col, w in zip("ABCDEFGH", [26, 11, 44, 16, 70, 11, 16, 30]):
        ws2.column_dimensions[col].width = w
    for r in range(2, n2 + 2):
        ws2.cell(row=r, column=5).alignment = WRAP; ws2.cell(row=r, column=3).alignment = WRAP

    # ---- Hoja 3: Plan de Gobierno (Keiko) ↔ Comisiones / Acuerdo Nacional
    n3 = 0
    try:
        keiko = L("keiko_articulacion.json")["propuestas"]
        ws3 = wb.create_sheet("Plan de Gobierno")
        cols3 = ["Pilar", "Propuesta (plan de gobierno)", "Resumen", "Se articula con", "Tipo (Comisión/AN)",
                 "Tipo propuesto (IA)", "Justificación", "¿Correcto?", "Observación"]
        ws3.append(cols3)
        for p in keiko:
            for c in p.get("comisiones", []):
                ws3.append([p["pilar"], p["titulo"], p.get("resumen", ""), "Comisión: " + c.get("comision_nombre", c.get("id", "")),
                            "Comisión", c["tipo"], c.get("justificacion", ""), "", ""]); n3 += 1
            for x in p.get("acuerdo_nacional", []):
                ws3.append([p["pilar"], p["titulo"], p.get("resumen", ""), "Política AN: P%s %s" % (x["politica"], x.get("politica_nombre", "")),
                            "Acuerdo Nacional", x["tipo"], x.get("justificacion", ""), "", ""]); n3 += 1
        style_header(ws3, len(cols3)); add_validation(ws3, "H", "F", n3)
        for col, w in zip("ABCDEFGHI", [12, 34, 46, 40, 16, 16, 60, 11, 30]):
            ws3.column_dimensions[col].width = w
        for r in range(2, n3 + 2):
            for cc in (3, 7): ws3.cell(row=r, column=cc).alignment = WRAP
    except Exception:
        pass

    # ---- Hoja Resumen
    wsr = wb.create_sheet("Resumen", 0)
    wsr.append(["Matriz de articulación — Plan Perú 2050 (propuesta con IA, a validar)"])
    wsr["A1"].font = Font(bold=True, size=13)
    wsr.append([])
    wsr.append(["Enlaces Comisiones ↔ Políticas de Estado (Acuerdo Nacional)", n])
    wsr.append(["Enlaces Comisiones ↔ Objetivos del Acuerdo Nacional (nivel-3)", nobj])
    wsr.append(["Enlaces Comisiones ↔ Programas Presupuestales (MEF)", n2])
    wsr.append(["Enlaces Plan de Gobierno ↔ Comisiones/AN", n3])
    wsr.append([])
    wsr.append(["Tipos de relación:"])
    wsr.append(["  igual/similar", "igualdad o similitud semántica directa"])
    wsr.append(["  desagregado", "la comisión/propuesta especifica algo contenido en la política/programa"])
    wsr.append(["  causal", "relación causa-efecto"])
    wsr.append([])
    wsr.append(["Para validar: usá las columnas '¿Correcto?' (Sí/No/Revisar), 'Tipo corregido' y 'Observación' en cada hoja."])
    wsr.column_dimensions["A"].width = 58; wsr.column_dimensions["B"].width = 60

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"✓ {OUT} — AN:{n} · PP:{n2} · Keiko:{n3}")


if __name__ == "__main__":
    main()
